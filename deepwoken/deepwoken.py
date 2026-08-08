from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import discord
from openpyxl import load_workbook
from redbot.core import commands

STAT_ALIASES = {
    "light": "LHT", "lht": "LHT", "medium": "MED", "med": "MED", "heavy": "HVY", "hvy": "HVY",
    "flame": "FIR", "fire": "FIR", "fir": "FIR", "frost": "ICE", "ice": "ICE",
    "lightning": "LTN", "thunder": "LTN", "ltn": "LTN", "wind": "WND", "gale": "WND", "wnd": "WND",
    "shadow": "SDW", "sdw": "SDW", "blood": "BLD", "bloodrend": "BLD", "bld": "BLD",
    "metal": "MTL", "mtl": "MTL", "strength": "STR", "str": "STR", "fortitude": "FTD", "ftd": "FTD",
    "agility": "AGI", "agi": "AGI", "intelligence": "INT", "int": "INT", "charisma": "CHA", "cha": "CHA",
    "willpower": "WLL", "will": "WLL", "wll": "WLL", "mind": "MND", "mnd": "MND", "body": "BDY", "bdy": "BDY",
}

RANKING_EXCLUSIONS = {
    "Ebonshard Lexicon", "The Rock", "The Endless Wave", "Unsung Scythern", "Worldpainter Brush",
    "Keyblade", "Soulshot", "Metal Greatsword", "Prototype Railblade", "Par's Glaive", "Saintsblade",
    "Ferractine", "Formless Shard", "Handcuffs",
}


class Deepwoken(commands.Cog):
    """Deepwoken weapon lookup and multi-stat sustained-DPS comparison."""

    def __init__(self, bot):
        self.bot = bot
        self.workbook_path = Path(__file__).parent / "data" / "weapons.xlsx"
        self.weapons = self._load_weapons()

    def _load_weapons(self) -> list[dict[str, Any]]:
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Missing bundled workbook: {self.workbook_path}")
        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        sheet = workbook["All Weapons"]
        headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(max_row=1))]
        weapons, seen = [], set()
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            name = str(row.get("Name") or "").strip()
            if not name:
                continue
            key = (name.casefold(), str(row.get("Requirements") or ""), str(row.get("Base Damage") or ""), str(row.get("Scaling") or ""), str(row.get("Swing Speed") or ""))
            if key not in seen:
                seen.add(key)
                weapons.append(row)
        workbook.close()
        return weapons

    @staticmethod
    def _number(value: Any) -> float | None:
        match = re.search(r"-?\d+(?:\.\d+)?", str(value or ""))
        return float(match.group()) if match else None

    @staticmethod
    def _requirements(value: Any) -> dict[str, int]:
        text = str(value or "").upper().replace(" OR ", " ")
        return {stat: int(amount) for amount, stat in re.findall(r"(\d+)\s*([A-Z]{2,4})(?=\s|\d|$)", text)}

    @staticmethod
    def _scaling(value: Any) -> dict[str, float]:
        return {stat: float(amount) for stat, amount in re.findall(r"([A-Z]{2,4})\s*:\s*(\d+(?:\.\d+)?)", str(value or "").upper())}

    def _parse_stat_query(self, args: tuple[str, ...]) -> tuple[str, dict[str, int], int] | None:
        if len(args) < 2 or args[0].casefold() not in STAT_ALIASES:
            return None
        primary = STAT_ALIASES[args[0].casefold()]

        # Backward-compatible: >dwweapon heavy 100 6
        if len(args) == 3 and args[1].isdigit() and args[2].isdigit():
            stats, proficiency = {primary: int(args[1])}, int(args[2])
            return (primary, stats, proficiency) if 0 <= stats[primary] <= 100 and 0 <= proficiency <= 6 else None

        stats, index, proficiency = {}, 0, 0
        while index < len(args):
            token = args[index].casefold()
            if token in {"prof", "proficiency"}:
                if index + 1 >= len(args) or not args[index + 1].isdigit():
                    return None
                proficiency = int(args[index + 1])
                index += 2
                continue
            if token not in STAT_ALIASES or index + 1 >= len(args) or not args[index + 1].isdigit():
                return None
            stats[STAT_ALIASES[token]] = int(args[index + 1])
            index += 2

        if primary not in stats or not all(0 <= value <= 100 for value in stats.values()) or not 0 <= proficiency <= 6:
            return None
        return primary, stats, proficiency

    def _meets_requirements(self, row: dict[str, Any], stats: dict[str, int]) -> bool:
        for stat, needed in self._requirements(row.get("Requirements")).items():
            # Power level is not an attribute investment; assume the build has sufficient Power.
            if stat == "LVL":
                continue
            if stats.get(stat, 0) < needed:
                return False
        return True

    def _damage(self, row: dict[str, Any], stats: dict[str, int], proficiency: int) -> float | None:
        base = self._number(row.get("Base Damage"))
        if base is None:
            return None
        multiplier = 1 + proficiency * 0.065
        damage = base
        for stat, scale in self._scaling(row.get("Scaling")).items():
            # Unspecified stats are zero, not silently set to weapon requirements.
            damage += 0.00075 * base * scale * stats.get(stat, 0) * multiplier
        return damage

    def _sustained_dps(self, row: dict[str, Any], stats: dict[str, int], proficiency: int) -> float | None:
        damage = self._damage(row, stats, proficiency)
        speed = self._number(row.get("Swing Speed"))
        if damage is None or speed is None or speed <= 0:
            return None
        endlag = self._number(row.get("Endlag")) or 0.0
        return damage / ((1 / (speed * 2)) + endlag)

    def _is_rankable(self, row: dict[str, Any]) -> bool:
        name = str(row.get("Name") or "").strip()
        weapon_class = str(row.get("Weapon Class") or "")
        if name in RANKING_EXCLUSIONS:
            return False
        if weapon_class in {"Special / Other", "Elemental", "Fighting Style"}:
            return False
        return not any(value > 100 for value in self._requirements(row.get("Requirements")).values())

    def _find_weapon(self, query: str) -> tuple[dict[str, Any] | None, list[dict[str, Any]]]:
        query = query.casefold().strip()
        exact = [row for row in self.weapons if str(row.get("Name") or "").casefold() == query]
        if exact:
            return exact[0], []
        matches = [row for row in self.weapons if query in str(row.get("Name") or "").casefold()]
        return (matches[0], []) if len(matches) == 1 else (None, matches)

    @commands.command(name="dwweapon", aliases=["dw", "weapon"])
    async def dwweapon(self, ctx: commands.Context, *args: str):
        """[p]dwweapon <name> or [p]dwweapon heavy 25 flame 75 light 10 prof 6."""
        if not args:
            await ctx.send("Use `[p]dwweapon <name>` or `[p]dwweapon heavy 25 flame 75 light 10 prof 6`.")
            return
        parsed = self._parse_stat_query(args)
        if parsed:
            primary, stats, proficiency = parsed
            await self._compare(ctx, primary, stats, proficiency)
        else:
            await self._lookup(ctx, " ".join(args))

    async def _compare(self, ctx: commands.Context, primary: str, stats: dict[str, int], proficiency: int):
        ranked = []
        for row in self.weapons:
            if not self._is_rankable(row):
                continue
            if primary not in self._scaling(row.get("Scaling")):
                continue
            if not self._meets_requirements(row, stats):
                continue
            damage = self._damage(row, stats, proficiency)
            dps = self._sustained_dps(row, stats, proficiency)
            if damage is not None and dps is not None:
                ranked.append((dps, damage, row))

        if not ranked:
            await ctx.send("No rankable weapons match those stats.")
            return

        ranked.sort(key=lambda item: item[0], reverse=True)
        stat_text = " | ".join(f"{stat} {value}" for stat, value in stats.items())
        lines = []
        for index, (dps, damage, row) in enumerate(ranked[:15], 1):
            label = str(row.get("Weapon Type") or "Unknown")
            if str(row.get("Weapon Class") or "") == "Crazy Slots":
                label += " | Crazy Slots"
            lines.append(f"`{index:>2}.` **{row['Name']}** ({label})\n`DPS:` {dps:.2f} | `M1:` {damage:.2f}")

        embed = discord.Embed(title=f"Weapon ranking | {stat_text} | Prof {proficiency}", description="\n".join(lines), colour=discord.Colour.blurple())
        embed.set_footer(text="Only weapons whose full stat requirements are met are shown. Crazy Slots and matching hybrids are included. DPS includes listed Endlag; bleed, crits, procs, enchants, talents, PEN, and resistance are excluded.")
        await ctx.send(embed=embed)

    async def _lookup(self, ctx: commands.Context, query: str):
        row, matches = self._find_weapon(query)
        if row is None:
            await ctx.send("Multiple matches: " + ", ".join(str(item["Name"]) for item in matches[:10]) if matches else "Weapon not found.")
            return
        embed = discord.Embed(title=str(row.get("Name") or "Unknown weapon"), colour=discord.Colour.blurple())
        embed.add_field(name="Class / Type", value=f"{row.get('Weapon Class') or '?'} / {row.get('Weapon Type') or '?'}", inline=False)
        for label, key in [("Requirements", "Requirements"), ("Base Damage", "Base Damage"), ("Scaled Damage", "Scaled Damage"), ("Scaling", "Scaling"), ("Armor Penetration", "Armor Penetration"), ("Chip Damage", "Chip Damage"), ("Posture Damage", "Posture Damage"), ("Range", "Range"), ("Swing Speed", "Swing Speed"), ("Endlag", "Endlag"), ("Tags", "Tags")]:
            embed.add_field(name=label, value=str(row.get(key) or "N/A"), inline=label != "Requirements")
        embed.set_footer(text="Example: [p]dwweapon heavy 25 flame 75 light 10 prof 6")
        await ctx.send(embed=embed)

    @commands.command(name="dwreload")
    @commands.is_owner()
    async def dwreload(self, ctx: commands.Context):
        self.weapons = self._load_weapons()
        await ctx.send(f"Reloaded {len(self.weapons)} unique weapon rows.")
